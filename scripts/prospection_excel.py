"""Ecrit la base de prospects en classeur Excel exploitable a la main.

Separe de `prospection.py` pour une raison simple : la mise en forme d'un
classeur est le seul endroit de l'outil qui depende d'une bibliotheque tierce
(openpyxl). La garder a part permet aux trois autres sous-commandes de tourner
sur une installation Python nue.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Ordre d'affichage : ce qui sert a DECIDER en premier, ce qui ne sert qu'a
# tracer la provenance a droite.
COLONNES = [
    ("organisation", "Organisation", 36),
    ("secteur", "Secteur", 24),
    ("commune", "Commune", 18),
    ("code_postal", "CP", 8),
    ("a_deja_un_site", "Site ?", 12),
    ("site_web", "Site trouvé", 28),
    ("proposition", "Ce qu'on peut proposer", 58),
    ("email_generique", "Email générique", 26),
    ("statut", "Statut", 16),
    ("note", "Note", 28),
    ("adresse", "Adresse", 32),
    ("siren", "SIREN", 13),
    ("naf", "NAF", 9),
    ("date_creation", "Créée le", 12),
    ("effectif", "Effectif", 9),
    ("nature_juridique", "Nature jur.", 11),
    ("statut_diffusion", "Diffusion", 10),
    ("source", "Source", 32),
]

A_REMPLIR = {"email_generique", "statut", "note"}

ENCRE = "1B1F27"
ACCENT = "5B5BD6"
SURFACE = "F7F8FA"
A_FAIRE = "FFF4E5"
LIGNE = "E2E5EA"

_fin = Side(style="thin", color=LIGNE)
BORDURE = Border(left=_fin, right=_fin, top=_fin, bottom=_fin)


def _feuille_prospects(ws, lignes: list[dict]) -> None:
    ws["A1"] = lignes[0].get("secteur") or "Prospects"
    ws["A1"].font = Font(size=16, bold=True, color=ENCRE)
    confirmes = sum(1 for l in lignes if (l.get("a_deja_un_site") or "") == "oui")
    ws["A2"] = (
        f"{len(lignes)} structures, toutes diffusibles au sens Sirene. "
        f"{confirmes} ont un site confirmé, {len(lignes) - confirmes} sont à vérifier. "
        "Aucune adresse email : l'annuaire public n'en publie pas."
    )
    ws["A2"].font = Font(size=10, color="5B6472")
    ws.append([])

    entete = 4
    for i, (_, libelle, largeur) in enumerate(COLONNES, start=1):
        c = ws.cell(row=entete, column=i, value=libelle)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = BORDURE
        ws.column_dimensions[get_column_letter(i)].width = largeur
    ws.row_dimensions[entete].height = 26

    enveloppe = {"adresse", "note", "organisation", "proposition"}
    for n, ligne in enumerate(lignes, start=entete + 1):
        for i, (cle, _, _) in enumerate(COLONNES, start=1):
            c = ws.cell(row=n, column=i, value=ligne.get(cle, ""))
            c.border = BORDURE
            c.alignment = Alignment(vertical="top", wrap_text=(cle in enveloppe))
            c.font = Font(size=10)
            if cle in A_REMPLIR:
                c.fill = PatternFill("solid", fgColor=A_FAIRE)
            elif n % 2 == 0:
                c.fill = PatternFill("solid", fgColor=SURFACE)
        ws.row_dimensions[n].height = 46

    ws.freeze_panes = ws.cell(row=entete + 1, column=1)
    derniere = get_column_letter(len(COLONNES))
    ws.auto_filter.ref = f"A{entete}:{derniere}{entete + len(lignes)}"

    def valide(cle: str, options: str) -> None:
        i = [c[0] for c in COLONNES].index(cle) + 1
        lettre = get_column_letter(i)
        dv = DataValidation(type="list", formula1=options, allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{lettre}{entete + 1}:{lettre}{entete + len(lignes)}")

    valide("a_deja_un_site", '"oui,à vérifier,non (vérifié à la main)"')
    valide(
        "statut",
        '"à qualifier,écarté,à contacter,contacté,relancé,répondu,rendez-vous,client,refus"',
    )


def _feuille_methode(ws, lignes: list[dict]) -> None:
    ws.column_dimensions["A"].width = 112

    def titre(t: str) -> None:
        ws.append([t])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12, color=ENCRE)
        ws.row_dimensions[ws.max_row].height = 22

    def para(t: str) -> None:
        ws.append([t])
        c = ws.cell(row=ws.max_row, column=1)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.font = Font(size=10)
        ws.row_dimensions[ws.max_row].height = max(15, 14 * (len(t) // 108 + 1))

    titre("La colonne « Site ? » ne dit jamais « non »")
    para(
        "Elle vaut « oui » quand un site a été trouvé ET que la page parle bien de "
        "l'entreprise, « à vérifier » sinon. On peut prouver qu'un site existe, jamais "
        "qu'il n'existe pas : l'entreprise peut être sur une page Facebook, un annuaire, "
        "ou un domaine que rien ne permet de deviner."
    )
    para(
        "Une première version se contentait de deviner un domaine, sans lire la page. Elle "
        "rendait 8 sites sur 10, dont les.fr, union.fr et soc.fr : des domaines "
        "appartenant à des tiers. Le faux positif est le pire des deux, parce qu'il marque "
        "un vrai prospect comme déjà équipé et le SORT de la liste."
    )
    ws.append([])

    titre("D'où vient cette liste")
    para(
        "API Recherche d'entreprises (recherche-entreprises.api.gouv.fr), service public "
        "gratuit, adossé aux bases Sirene et RNE. Générée par scripts/prospection.py : la "
        "liste est reproductible et sa provenance traçable, ce qu'exige l'article 14 du "
        "RGPD pour une donnée qu'on n'a pas reçue de la personne."
    )
    ws.append([])

    titre("Ce qui a été volontairement écarté")
    para(
        "Les structures opposées à la diffusion de leurs données Sirene. Plus d'un million "
        "d'établissements l'ont demandé à l'INSEE : c'est une opposition DÉJÀ EXERCÉE, et "
        "la respecter passe avant tout le reste."
    )
    para(
        "Les entreprises individuelles, par prudence et non par obligation. La CNIL admet "
        "l'intérêt légitime dès lors que l'objet du message relève de la profession du "
        "destinataire, mais sa dérogation explicite ne vise que les personnes morales."
    )
    para("Les structures de 20 salariés et plus : elles ont déjà un prestataire.")
    ws.append([])

    titre("La règle, en une phrase")
    para(
        "J'écris sans consentement préalable à un professionnel, sur une adresse "
        "professionnelle, pour une offre qui relève de sa FONCTION, en disant qui je suis, "
        "d'où je tiens son adresse et comment ne plus jamais recevoir de message."
    )
    para(
        "Le critère porte sur la fonction du destinataire, pas sur le secteur. Écrire au "
        "gérant d'un restaurant pour un site : conforme. Écrire à son cuisinier : non."
    )
    ws.append([])

    titre("Les adresses que la base accepte")
    para(
        "contact, info, bonjour, hello, accueil, direction, secretariat, commercial, admin. "
        "Une adresse prenom.nom@ est refusée par la contrainte de la base, ce n'est pas un "
        "réglage contournable."
    )
    ws.append([])

    titre("Ce qu'il ne faut pas faire")
    para(
        "Ne rien envoyer par le moteur tant qu'aucune formulation n'a reçu de réponse "
        "humaine. Automatiser un message qui ne marche pas ne fait que l'amplifier. Les "
        "trente premiers s'écrivent à la main."
    )
    para(
        "Ne pas recopier ce fichier dans le dépôt : il porte des données relatives à des "
        "personnes, et le dépôt est public. L'outil refuse d'ailleurs d'y écrire."
    )
    ws.append([])

    titre("Répartition")
    for cle, libelle in (("commune", "Communes"), ("secteur", "Secteurs")):
        compte: dict[str, int] = {}
        for l in lignes:
            v = l.get(cle) or "?"
            compte[v] = compte.get(v, 0) + 1
        para(libelle + " :")
        for nom, n in sorted(compte.items(), key=lambda x: -x[1])[:20]:
            para(f"    {nom} : {n}")
        ws.append([])


def _nom_onglet(secteur: str) -> str:
    """Excel refuse plus de 31 caracteres, et certains caracteres dans un nom
    d'onglet : deux-points, barres obliques, point d'interrogation, asterisque
    et crochets."""
    nom = "".join(c for c in secteur if c not in set(":\\/?*[]"))
    return (nom or "Autres")[:31]


def ecris_classeur(lignes: list[dict], sortie: str) -> None:
    """Un onglet par niche, plus la synthese et la methode.

    POURQUOI UN ONGLET PAR NICHE. Un message qui parle du metier de celui qui le
    lit n'a rien a voir avec un message generique, et le tri par secteur est ce
    qui rend ce travail possible : on ecrit une fois par niche, pas une fois par
    prospect. Une liste unique de plusieurs milliers de lignes melangees se
    parcourt, elle ne se travaille pas.
    """
    groupes: dict[str, list[dict]] = {}
    for l in lignes:
        groupes.setdefault(l.get("secteur") or "Autres", []).append(l)

    classeur = Workbook()

    # Synthese en premier : c'est la page qu'on ouvre pour decider par ou
    # commencer, pas pour lire des lignes.
    ws = classeur.active
    ws.title = "Synthèse"
    ws.column_dimensions["A"].width = 52
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18

    ws["A1"] = "Prospects BLF Lab's"
    ws["A1"].font = Font(size=16, bold=True, color=ENCRE)
    ws["A2"] = (
        f"{len(lignes)} structures réparties en {len(groupes)} niches, une par onglet. "
        "Toutes diffusibles au sens Sirene."
    )
    ws["A2"].font = Font(size=10, color="5B6472")

    for i, libelle in enumerate(
        ["Niche", "Structures", "Site confirmé", "Site à vérifier"], start=1
    ):
        c = ws.cell(row=4, column=i, value=libelle)
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.border = BORDURE

    n = 5
    for secteur, groupe in sorted(groupes.items(), key=lambda x: -len(x[1])):
        avec = sum(1 for l in groupe if (l.get("a_deja_un_site") or "") == "oui")
        for i, v in enumerate(
            [secteur, len(groupe), avec, len(groupe) - avec], start=1
        ):
            c = ws.cell(row=n, column=i, value=v)
            c.border = BORDURE
            c.font = Font(size=10)
            c.alignment = Alignment(wrap_text=(i == 1), vertical="top")
        n += 1

    for secteur, groupe in sorted(groupes.items(), key=lambda x: -len(x[1])):
        _feuille_prospects(classeur.create_sheet(_nom_onglet(secteur)), groupe)

    _feuille_methode(classeur.create_sheet("Méthode et règles"), lignes)

    chemin = Path(sortie)
    chemin.parent.mkdir(parents=True, exist_ok=True)
    classeur.save(chemin)
    print(f"{len(lignes)} prospects, {len(groupes)} onglets, ecrits dans {chemin}")
